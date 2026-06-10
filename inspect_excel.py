import pandas as pd
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('CPF.xlsx')

# Get sheets
ws1 = wb['CPF TB']
ws2 = wb['Retirement Dashboard']

print('CPF TB Key Data:')
for row in ws1.iter_rows(values_only=True):
    print(row[:10])

print('\nRetirement Dashboard Key Data:')
for row in ws2.iter_rows(values_only=True):
    print(row[:10])

# Get formulas
print('\nCPF TB Formulas:')
for row in ws1.iter_rows():
    for cell in row:
        if cell.data_type == 'f' and cell.value:
            print(f"Cell {cell.coordinate}: {cell.value}")

print('\nRetirement Dashboard Formulas:')
for row in ws2.iter_rows():
    for cell in row:
        if cell.data_type == 'f' and cell.value:
            print(f"Cell {cell.coordinate}: {cell.value}")